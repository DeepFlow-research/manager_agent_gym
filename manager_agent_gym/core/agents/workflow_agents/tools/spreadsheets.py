"""Spreadsheet tools (Excel, CSV, Charts) - two-layer architecture.

Layer 1: Core functions (_*) - pure business logic, testable, returns typed results
Layer 2: OpenAI tool wrappers - thin adapters for OpenAI SDK, handle JSON serialization
"""

import json
import re
from pathlib import Path
from typing import TYPE_CHECKING, Any

import openpyxl
import pandas as pd
from agents import Tool, function_tool
from openpyxl.chart import (
    AreaChart,
    BarChart,
    LineChart,
    PieChart,
    Reference,
    ScatterChart,
)
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from pydantic import BaseModel, Field

if TYPE_CHECKING:
    from manager_agent_gym.core.workflow.resource_storage import ResourceFileManager


# ============================================================================
# PYDANTIC MODELS (Shared)
# ============================================================================


class ExcelData(BaseModel):
    """Data structure for Excel table data."""

    headers: list[str] = Field(
        default_factory=list, description="Column headers for the Excel sheet"
    )
    rows: list[list] = Field(
        ..., description="Data rows as list of lists (each inner list is one row)"
    )


class CSVData(BaseModel):
    """Data structure for CSV data."""

    columns: list[str] = Field(..., description="Column names for the CSV")
    data: list[list[Any]] = Field(
        ..., description="Data rows as list of lists (each inner list is one row)"
    )


# ============================================================================
# LAYER 1: EXCEL OPERATIONS (Core Business Logic)
# ============================================================================


async def _read_excel(file_path: str, sheet_name: str | None = None) -> dict[str, Any]:
    """Read data from Excel file."""
    try:
        file_path_obj = Path(file_path)
        if not file_path_obj.exists():
            return {"success": False, "error": "Excel file not found"}

        workbook = openpyxl.load_workbook(file_path, data_only=True)

        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                return {
                    "success": False,
                    "error": f"Sheet '{sheet_name}' not found",
                }
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        if sheet is None:
            return {"success": False, "error": "No active worksheet found"}

        data = []
        for row in sheet.iter_rows(values_only=True):
            row_data = [
                ""
                if cell is None
                else str(cell)
                if not isinstance(cell, (str, int, float, bool))
                else cell
                for cell in row
            ]
            data.append(row_data)

        return {
            "success": True,
            "sheet_name": sheet.title,
            "num_rows": len(data),
            "num_cols": len(data[0]) if data else 0,
            "data": data,
        }

    except Exception as e:
        return {"success": False, "error": f"Error reading Excel: {str(e)}"}


async def _create_excel(
    data: ExcelData, output_path: str, sheet_name: str = "Sheet1"
) -> dict[str, Any]:
    """Create Excel file from data."""
    try:
        output_path_obj = Path(output_path)
        output_path_obj.parent.mkdir(parents=True, exist_ok=True)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        if sheet is None:
            return {"success": False, "error": "Failed to create worksheet"}
        sheet.title = sheet_name

        has_headers = bool(data.headers)
        if has_headers:
            sheet.append(data.headers)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

        # Add data rows (can be empty)
        for row_data in data.rows:
            sheet.append(row_data)

        for column in list(sheet.columns):
            max_length = 0
            first_cell = column[0]
            try:
                if hasattr(first_cell, "column") and first_cell.column is not None:
                    column_letter = get_column_letter(first_cell.column)
                else:
                    continue
            except AttributeError:
                continue

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(str(output_path_obj))
        file_size = output_path_obj.stat().st_size

        return {
            "success": True,
            "file_path": str(output_path_obj.absolute()),
            "file_name": output_path_obj.name,
            "size_bytes": file_size,
            "num_rows": len(data.rows),
            "num_columns": len(data.headers) if data.headers else 0,
        }

    except Exception as e:
        return {"success": False, "error": f"Error creating Excel: {str(e)}"}


async def _add_excel_sheet(
    file_path: str, sheet_name: str, data: ExcelData
) -> dict[str, Any]:
    """Add new sheet to existing Excel file."""
    try:
        file_path_obj = Path(file_path)
        if not file_path_obj.exists():
            return {"success": False, "error": "Excel file not found"}

        workbook = openpyxl.load_workbook(file_path)

        if sheet_name in workbook.sheetnames:
            return {"success": False, "error": f"Sheet '{sheet_name}' already exists"}

        sheet = workbook.create_sheet(title=sheet_name)

        if data.headers:
            sheet.append(data.headers)
            header_font = Font(bold=True)
            for cell in sheet[1]:
                cell.font = header_font

        if data.rows:
            for row_data in data.rows:
                sheet.append(row_data)

        workbook.save(str(file_path_obj))

        return {"success": True, "output_path": str(file_path_obj.absolute())}

    except Exception as e:
        return {"success": False, "error": f"Error adding sheet: {str(e)}"}


async def _format_excel_cells(
    file_path: str,
    sheet_name: str,
    cell_range: str,
    format_type: str,
    value: str | None = None,
) -> dict[str, Any]:
    """Apply formatting to Excel cells."""
    try:
        file_path_obj = Path(file_path)
        if not file_path_obj.exists():
            return {"success": False, "error": "Excel file not found"}

        workbook = openpyxl.load_workbook(file_path)

        if sheet_name not in workbook.sheetnames:
            return {"success": False, "error": f"Sheet '{sheet_name}' not found"}

        sheet = workbook[sheet_name]

        for row in sheet[cell_range]:
            for cell in row if isinstance(row, tuple) else [row]:
                if format_type == "number":
                    cell.number_format = value or "#,##0.00"
                elif format_type == "currency":
                    cell.number_format = value or "$#,##0.00"
                elif format_type == "percent":
                    cell.number_format = value or "0.00%"
                elif format_type == "date":
                    cell.number_format = value or "yyyy-mm-dd"
                elif format_type == "bold":
                    cell.font = Font(bold=True)
                elif format_type == "color":
                    if value:
                        # Convert RGB hex (#RRGGBB) to aRGB hex (AARRGGBB) if needed
                        color_value = value.lstrip("#")
                        if len(color_value) == 6:
                            # Add full opacity (FF) prefix for RGB colors
                            color_value = "FF" + color_value
                        elif len(color_value) != 8:
                            return {
                                "success": False,
                                "error": f"Invalid color format: {value}. Use #RRGGBB or #AARRGGBB",
                            }

                        cell.fill = PatternFill(
                            start_color=color_value,
                            end_color=color_value,
                            fill_type="solid",
                        )

        workbook.save(str(file_path_obj))

        return {"success": True}

    except Exception as e:
        return {"success": False, "error": f"Error formatting cells: {str(e)}"}


async def _get_excel_info(file_path: str) -> dict[str, Any]:
    """Get information about Excel file."""
    try:
        file_path_obj = Path(file_path)
        if not file_path_obj.exists():
            return {"success": False, "error": "Excel file not found"}

        workbook = openpyxl.load_workbook(file_path, data_only=True)

        sheets_info = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheets_info.append(
                {
                    "name": sheet_name,
                    "rows": sheet.max_row,
                    "columns": sheet.max_column,
                }
            )

        return {
            "success": True,
            "file_path": str(file_path),
            "num_sheets": len(workbook.sheetnames),
            "sheets": sheets_info,
        }

    except Exception as e:
        return {"success": False, "error": f"Error getting Excel info: {str(e)}"}


# ============================================================================
# LAYER 1: CSV OPERATIONS
# ============================================================================


async def _read_csv(file_path: str, max_rows: int | None = None) -> dict[str, Any]:
    """Read CSV file."""
    try:
        file_path_obj = Path(file_path)
        if not file_path_obj.exists():
            return {"success": False, "error": "CSV file not found"}

        if max_rows:
            df = pd.read_csv(file_path, nrows=max_rows)
        else:
            df = pd.read_csv(file_path)

        result = {
            "num_rows": len(df),
            "num_columns": len(df.columns),
            "columns": df.columns.tolist(),
            "data": df.to_dict(orient="records"),
        }

        return {"success": True, **result}

    except Exception as e:
        return {"success": False, "error": f"Error reading CSV: {str(e)}"}


async def _write_csv(
    data: CSVData, output_path: str, include_index: bool = False
) -> dict[str, Any]:
    """Write CSV file."""
    try:
        output_path_obj = Path(output_path)
        output_path_obj.parent.mkdir(parents=True, exist_ok=True)

        if data.data:
            df = pd.DataFrame(data.data)
            if data.columns:
                df.columns = pd.Index(data.columns)
        elif data.columns:
            df = pd.DataFrame(columns=pd.Index(data.columns))
        else:
            return {"success": False, "error": "Data must contain data or columns"}

        df.to_csv(str(output_path_obj), index=include_index)

        return {
            "success": True,
            "output_path": str(output_path_obj.absolute()),
            "num_rows": len(df),
            "num_columns": len(df.columns),
        }

    except Exception as e:
        return {"success": False, "error": f"Error writing CSV: {str(e)}"}


async def _analyze_csv(file_path: str) -> dict[str, Any]:
    """Analyze CSV file."""
    try:
        file_path_obj = Path(file_path)
        if not file_path_obj.exists():
            return {"success": False, "error": "CSV file not found"}

        df = pd.read_csv(file_path)

        info = {
            "num_rows": len(df),
            "num_columns": len(df.columns),
            "columns": df.columns.tolist(),
            "dtypes": {col: str(dtype) for col, dtype in df.dtypes.items()},
            "missing_values": df.isnull().sum().to_dict(),
        }

        numeric_cols = df.select_dtypes(include=["number"]).columns.tolist()
        if numeric_cols:
            info["numeric_summary"] = {
                col: {
                    "mean": float(df[col].mean()),
                    "min": float(df[col].min()),
                    "max": float(df[col].max()),
                    "std": float(df[col].std()),
                }
                for col in numeric_cols
            }

        return {"success": True, **info}

    except Exception as e:
        return {"success": False, "error": f"Error analyzing CSV: {str(e)}"}


# ============================================================================
# LAYER 1: EXCEL CHART OPERATIONS
# ============================================================================


async def _add_excel_chart(
    file_path: str,
    sheet_name: str,
    chart_type: str,
    data_range: str,
    chart_title: str,
    position: str = "E2",
) -> dict[str, Any]:
    """Add chart to Excel file."""
    try:
        file_path_obj = Path(file_path)
        if not file_path_obj.exists():
            return {"success": False, "error": "Excel file not found"}

        workbook = openpyxl.load_workbook(file_path)

        if sheet_name not in workbook.sheetnames:
            return {"success": False, "error": f"Sheet '{sheet_name}' not found"}

        sheet = workbook[sheet_name]

        if sheet is None:
            return {"success": False, "error": "Failed to load sheet"}

        if chart_type.lower() == "bar":
            chart = BarChart()
        elif chart_type.lower() == "line":
            chart = LineChart()
        elif chart_type.lower() == "pie":
            chart = PieChart()
        elif chart_type.lower() == "area":
            chart = AreaChart()
        elif chart_type.lower() == "scatter":
            chart = ScatterChart()
        else:
            return {
                "success": False,
                "error": f"Unknown chart type: {chart_type}",
            }

        chart.title = chart_title

        try:
            parts = data_range.split(":")
            if len(parts) != 2:
                return {"success": False, "error": "Invalid range format"}

            match_start = re.match(r"([A-Z]+)(\d+)", parts[0])
            match_end = re.match(r"([A-Z]+)(\d+)", parts[1])

            if not match_start or not match_end:
                return {"success": False, "error": "Invalid cell references"}

            min_col = column_index_from_string(match_start.group(1))
            min_row = int(match_start.group(2))
            max_col = column_index_from_string(match_end.group(1))
            max_row = int(match_end.group(2))

            data_ref = Reference(
                sheet,
                min_col=min_col,
                min_row=min_row,
                max_col=max_col,
                max_row=max_row,
            )
            chart.add_data(data_ref, titles_from_data=True)
        except Exception as e:
            return {"success": False, "error": f"Error parsing range: {str(e)}"}

        sheet.add_chart(chart, position)  # type: ignore[call-arg]

        workbook.save(str(file_path_obj))

        return {"success": True, "chart_type": chart_type}

    except Exception as e:
        return {"success": False, "error": f"Error adding chart: {str(e)}"}


# ============================================================================
# LAYER 2: OPENAI TOOL WRAPPERS
# ============================================================================


def create_spreadsheets_tools(resource_manager: "ResourceFileManager") -> list[Tool]:
    """Create spreadsheet tools for OpenAI SDK."""
    from manager_agent_gym.core.workflow.context import AgentExecutionContext
    from agents import RunContextWrapper

    @function_tool
    async def read_excel(file_path: str, sheet_name: str | None = None) -> str:
        """
        Read and extract data from Microsoft Excel files (.xlsx, .xls).

        This tool reads Excel workbooks and extracts all data from a specified sheet or
        the active sheet. All cell values are preserved including formulas (calculated
        values), numbers, text, and dates. Perfect for analyzing spreadsheets, extracting
        data, or processing Excel reports.

        Parameters:
            file_path (str):
                The path to the Excel file to read.
                Example: "/path/to/spreadsheet.xlsx" or "reports/data.xlsx"
            sheet_name (str | None):
                Optional. The name of the specific sheet to read. If None, reads the
                active (first) sheet. Default: None.

        Returns:
            str:
                JSON string containing:
                - success: Whether the operation succeeded
                - sheet_name: Name of the sheet that was read
                - num_rows: Total number of rows
                - num_cols: Total number of columns
                - data: 2D array of cell values (rows and columns)

        Usage:
            Call this tool when you need to read data from Excel files for analysis,
            reporting, or data processing. Use it to extract tables, reports, datasets,
            or any structured data stored in Excel format.
        """
        result = await _read_excel(file_path, sheet_name)
        return json.dumps(result, indent=2)

    @function_tool
    async def create_excel(
        ctx: RunContextWrapper[AgentExecutionContext],
        data: ExcelData,
        output_path: str,
        sheet_name: str = "Sheet1",
    ) -> str:
        """
        Create a professionally formatted Excel spreadsheet from structured data.

        This tool generates Excel files with automatic formatting including styled headers,
        auto-sized columns, and professional color schemes. Perfect for creating reports,
        data exports, or formatted spreadsheets from processed data.

        **✨ AUTOMATIC FILE TRACKING:**
        Files you create are **automatically tracked as resources**! You no longer need to:
        - Parse JSON responses
        - Manually construct Resource objects
        - Worry about file paths

        Just create the file and reference it in your notes - the system handles the rest.

        Parameters:
            data (ExcelData):
                The data structure containing:
                - headers: List of column header names (optional, can be empty)
                - rows: List of lists, where each inner list represents a row of data
                Example: ExcelData(headers=["Name", "Age"], rows=[["Alice", 30], ["Bob", 25]])
            output_path (str):
                The path where the Excel file should be saved.
                Example: "/path/to/output.xlsx"
            sheet_name (str):
                The name for the worksheet. Default: "Sheet1".

        Returns:
            str:
                Human-readable summary: "✅ Created Excel file: filename.xlsx (X rows, Y KB)"

        Usage:
            Use this tool to create Excel spreadsheets from data you've generated or
            processed. Perfect for creating reports, exporting analysis results, or
            generating formatted data files for sharing.
        """
        result = await _create_excel(data, output_path, sheet_name)

        # AUTO-REGISTER: Track created file as intermediary resource
        if result["success"]:
            try:
                if ctx.context:
                    from manager_agent_gym.schemas.domain.resource import Resource

                    ctx.context.register_created_resource(
                        Resource(
                            name=f"Generated: {result['file_name']}",
                            description="Auto-created by create_excel tool",
                            file_path=result["file_path"],
                            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            size_bytes=result["size_bytes"],
                            resource_role="intermediary",
                        )
                    )
            except Exception as e:
                from manager_agent_gym.core.common.logging import logger

                logger.warning(f"Failed to auto-register Excel resource: {e}")

            # Return human-readable summary
            return (
                f"✅ Created Excel file: {result['file_name']} "
                f"({result['num_rows']} rows, {result['size_bytes']} bytes)\n"
                f"Location: {result['file_path']}"
            )

        return f"❌ Error creating Excel: {result.get('error', 'Unknown error')}"

    @function_tool
    async def add_excel_sheet(file_path: str, sheet_name: str, data: ExcelData) -> str:
        """
        Add a new worksheet with data to an existing Excel workbook.

        This tool appends a new sheet to an existing Excel file with formatted headers
        and data. Useful for creating multi-sheet workbooks with different datasets or
        adding new data to existing reports.

        Parameters:
            file_path (str):
                The path to an existing Excel file where the sheet will be added.
                Example: "/path/to/workbook.xlsx"
            sheet_name (str):
                The name for the new sheet. Must be unique within the workbook.
                Example: "Q4 Results" or "Summary"
            data (ExcelData):
                The data structure containing headers and rows.
                Example: ExcelData(headers=["Product", "Sales"], rows=[["A", 100]])

        Returns:
            str:
                Success message confirming the sheet was added, or an error message if
                the operation fails (e.g., sheet name already exists).

        Usage:
            Call this tool when you need to add additional sheets to an existing Excel
            workbook. Common uses include adding summary sheets, different time periods,
            or separate datasets to a single workbook.
        """
        result = await _add_excel_sheet(file_path, sheet_name, data)
        if result["success"]:
            return f"Successfully added sheet '{sheet_name}' to {file_path}"
        return f"Error: {result.get('error')}"

    @function_tool
    async def format_excel_cells(
        file_path: str,
        sheet_name: str,
        cell_range: str,
        format_type: str,
        value: str | None = None,
    ) -> str:
        """
        Apply professional formatting to specific cell ranges in Excel spreadsheets.

        This tool formats cells with number formats, colors, or styles to make
        spreadsheets more readable and professional. Supports currency, percentages,
        dates, and visual styling.

        Parameters:
            file_path (str):
                The path to the Excel file to format.
                Example: "/path/to/spreadsheet.xlsx"
            sheet_name (str):
                The name of the sheet containing the cells to format.
            cell_range (str):
                The Excel range to format (e.g., "A1:B10", "C5", "D1:D100").
            format_type (str):
                The type of formatting to apply. Options:
                - "number": Number format (e.g., "1,234.56")
                - "currency": Currency format (e.g., "$1,234.56")
                - "percent": Percentage format (e.g., "12.34%")
                - "date": Date format (e.g., "2024-01-15")
                - "bold": Make text bold
                - "color": Apply background color
            value (str | None):
                Optional. Custom format string or color hex code (for color type).
                For colors, use standard hex format like "#FF0000" for red,
                "#00FF00" for green, etc. The tool automatically handles conversion
                to Excel's internal format.
                Default: None (uses standard formats).

        Returns:
            str:
                Success message confirming the formatting was applied, or an error
                message if the operation fails.

        Usage:
            Use this tool to format Excel cells after creating or modifying spreadsheets.
            Perfect for highlighting important data, formatting financial numbers, or
            making spreadsheets more visually appealing and readable.

            Example color usage:
            - Red: "#FF0000" or "#FFFF0000"
            - Green: "#00FF00"
            - Blue: "#0000FF"
            - Yellow: "#FFFF00"
        """
        result = await _format_excel_cells(
            file_path, sheet_name, cell_range, format_type, value
        )
        if result["success"]:
            return f"Successfully applied {format_type} formatting to {cell_range}"
        return f"Error: {result.get('error')}"

    @function_tool
    async def get_excel_info(file_path: str) -> str:
        """
        Retrieve metadata and structure information about an Excel workbook.

        This tool analyzes an Excel file and returns information about all sheets,
        including sheet names, row counts, and column counts. Useful for understanding
        workbook structure before reading or processing specific sheets.

        Parameters:
            file_path (str):
                The path to the Excel file to analyze.
                Example: "/path/to/workbook.xlsx"

        Returns:
            str:
                JSON string containing:
                - file_path: Path to the file
                - num_sheets: Total number of sheets in the workbook
                - sheets: Array of sheet information, each containing:
                  * name: Sheet name
                  * rows: Number of rows in the sheet
                  * columns: Number of columns in the sheet

        Usage:
            Call this tool before reading Excel files to understand their structure.
            Perfect for discovering what sheets exist in a workbook, determining which
            sheet to read, or getting an overview of the workbook's content.
        """
        result = await _get_excel_info(file_path)
        return json.dumps(result, indent=2)

    @function_tool
    async def read_csv(file_path: str, max_rows: int | None = None) -> str:
        """
        Read and parse CSV (Comma-Separated Values) files into structured data.

        This tool reads CSV files and converts them into structured JSON data with
        automatic type detection. You can optionally limit the number of rows read
        for large files. Perfect for processing data exports, logs, or tabular data.

        Parameters:
            file_path (str):
                The path to the CSV file to read.
                Example: "/path/to/data.csv" or "exports/report.csv"
            max_rows (int | None):
                Optional. Maximum number of rows to read. Useful for previewing large
                files or limiting memory usage. If None, reads all rows. Default: None.

        Returns:
            str:
                JSON string containing:
                - success: Whether the operation succeeded
                - num_rows: Number of rows read
                - num_columns: Number of columns
                - columns: List of column names
                - data: Array of row objects (each row as a dictionary)

        Usage:
            Use this tool to read CSV files for data analysis, processing, or conversion.
            Common uses include reading data exports, processing logs, analyzing datasets,
            or extracting information from CSV reports.
        """
        result = await _read_csv(file_path, max_rows)
        return json.dumps(result, indent=2, default=str)

    @function_tool
    async def write_csv(
        ctx: RunContextWrapper[AgentExecutionContext],
        data: CSVData,
        output_path: str,
        include_index: bool = False,
    ) -> str:
        """
        Create a CSV file from structured data for easy sharing and compatibility.

        This tool converts structured data into CSV format, which is universally
        compatible with spreadsheet applications, databases, and data analysis tools.
        Perfect for exporting data, creating data files, or sharing datasets.

        **✨ AUTOMATIC FILE TRACKING:**
        Files you create are **automatically tracked as resources**! Just create the file
        and reference it in your notes - the system handles resource management.

        Parameters:
            data (CSVData):
                The data structure containing:
                - columns: List of column names
                - data: List of lists, where each inner list represents a row
                Example: CSVData(columns=["Name", "Score"], data=[["Alice", 95], ["Bob", 87]])
            output_path (str):
                The path where the CSV file should be saved.
                Example: "/path/to/output.csv"
            include_index (bool):
                Whether to include a row number index column. Default: False.

        Returns:
            str:
                Human-readable summary: "✅ Created CSV: filename.csv (X rows)"

        Usage:
            Call this tool to export data to CSV format for sharing with others, importing
            into other systems, or creating data files that can be opened in any spreadsheet
            application. CSV is the most compatible data format.
        """
        result = await _write_csv(data, output_path, include_index)

        # AUTO-REGISTER: Track created file as intermediary resource
        if result["success"]:
            from pathlib import Path

            try:
                if ctx.context:
                    from manager_agent_gym.schemas.domain.resource import Resource

                    filename = Path(result["output_path"]).name
                    ctx.context.register_created_resource(
                        Resource(
                            name=f"Generated: {filename}",
                            description="Auto-created by write_csv tool",
                            file_path=result["output_path"],
                            mime_type="text/csv",
                            size_bytes=result["size_bytes"],
                            resource_role="intermediary",
                        )
                    )
            except Exception as e:
                from manager_agent_gym.core.common.logging import logger

                logger.warning(f"Failed to auto-register CSV resource: {e}")

            return f"✅ Created CSV: {Path(result['output_path']).name} ({result['num_rows']} rows, {result['size_bytes']} bytes)\nLocation: {result['output_path']}"

        return f"❌ Error creating CSV: {result.get('error', 'Unknown error')}"

    @function_tool
    async def analyze_csv(file_path: str) -> str:
        """
        Perform comprehensive statistical analysis on CSV file data.

        This tool analyzes CSV files and provides detailed statistics including data types,
        missing values, and numerical summaries (mean, min, max, standard deviation).
        Perfect for quickly understanding dataset characteristics and data quality.

        Parameters:
            file_path (str):
                The path to the CSV file to analyze.
                Example: "/path/to/dataset.csv"

        Returns:
            str:
                JSON string containing:
                - num_rows: Total number of data rows
                - num_columns: Total number of columns
                - columns: List of column names
                - dtypes: Data type for each column
                - missing_values: Count of missing values per column
                - numeric_summary: For numeric columns, statistics including mean, min, max, std

        Usage:
            Use this tool to quickly understand a CSV dataset before processing. Perfect
            for data quality checks, understanding data distributions, identifying missing
            values, or getting dataset statistics for reporting.
        """
        result = await _analyze_csv(file_path)
        return json.dumps(result, indent=2, default=str)

    @function_tool
    async def add_excel_chart(
        file_path: str,
        sheet_name: str,
        chart_type: str,
        data_range: str,
        chart_title: str,
        position: str = "E2",
    ) -> str:
        """
        Create and embed professional charts directly in Excel spreadsheets.

        This tool adds various types of charts (bar, line, pie, area, scatter) to Excel
        sheets to visualize data. The charts are embedded in the spreadsheet and can be
        customized with titles and positioning. Perfect for creating visual reports and
        data presentations.

        Parameters:
            file_path (str):
                The path to the Excel file where the chart will be added.
                Example: "/path/to/report.xlsx"
            sheet_name (str):
                The name of the sheet containing the data and where the chart will appear.
            chart_type (str):
                The type of chart to create. Options: "bar", "line", "pie", "area", "scatter".
            data_range (str):
                The Excel range containing the data to chart (e.g., "A1:B10").
                The first row is typically used as series labels.
            chart_title (str):
                The title to display on the chart.
                Example: "Monthly Sales Comparison"
            position (str):
                The cell position where the chart's top-left corner will be placed.
                Default: "E2".

        Returns:
            str:
                Success message confirming the chart was added, or an error message if
                the operation fails (e.g., invalid range, sheet not found).

        Usage:
            Call this tool to add visual representations of data to Excel reports. Perfect
            for creating dashboards, visualizing trends, comparing values, or making data
            more understandable through charts.
        """
        result = await _add_excel_chart(
            file_path, sheet_name, chart_type, data_range, chart_title, position
        )
        if result["success"]:
            return f"Successfully added {result['chart_type']} chart to {sheet_name}"
        return f"Error: {result.get('error')}"

    return [
        read_excel,
        create_excel,
        add_excel_sheet,
        format_excel_cells,
        get_excel_info,
        read_csv,
        write_csv,
        analyze_csv,
        add_excel_chart,
    ]
