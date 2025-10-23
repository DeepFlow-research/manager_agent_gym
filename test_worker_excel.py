"""
Minimum reproducible example: Single worker with Excel file task.

This tests whether the worker can successfully read an Excel file when provided as a resource.
"""

import asyncio
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from manager_agent_gym.schemas.domain.resource import Resource
from manager_agent_gym.schemas.domain.task import Task
from manager_agent_gym.schemas.domain.base import TaskStatus
from manager_agent_gym.schemas.agents import AIAgentConfig
from manager_agent_gym.core.agents.workflow_agents.workers.ai_agent import AIAgent
from manager_agent_gym.core.workflow.resource_storage import ResourceFileManager
from manager_agent_gym.core.agents.workflow_agents.tools.tool_factory import ToolFactory
from manager_agent_gym.core.communication.service import CommunicationService


async def test_worker_excel_file():
    """Test a single worker reading an Excel file."""

    print("=" * 80)
    print("MINIMAL TEST: Worker + Excel File")
    print("=" * 80)
    print()

    # 1. Create input Excel file
    print("Step 1: Creating Excel file...")
    temp_dir = Path(tempfile.mkdtemp(prefix="test_excel_"))
    input_excel_path = temp_dir / "customer_data.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create worksheet")
    ws.title = "Customers"

    # Headers
    headers = [
        "Customer_ID",
        "Tenure_Months",
        "Monthly_Spend",
        "Support_Tickets",
        "Usage_Score",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )

    # Sample data
    data = [
        [1001, 24, 89.99, 2, 85],
        [1002, 6, 45.50, 5, 40],
        [1003, 36, 120.00, 1, 95],
    ]
    for row in data:
        ws.append(row)

    wb.save(str(input_excel_path))
    print(f"  ✓ Created: {input_excel_path}")
    print(f"  ✓ File exists: {input_excel_path.exists()}")
    print()

    # 2. Create Resource
    print("Step 2: Creating Resource...")
    input_resource = Resource(
        name="Customer Data",
        description="Excel file containing customer metrics",
        file_path=str(input_excel_path.absolute()),
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        size_bytes=input_excel_path.stat().st_size,
        file_format_metadata={
            "sheet_names": ["Customers"],
            "num_rows": 3,
            "num_columns": 5,
        },
    )
    print(f"  ✓ Resource created: {input_resource.name}")
    print(f"  ✓ File path: {input_resource.file_path}")
    print()

    # 3. Create Task - FULL ML TASK
    print("Step 3: Creating Task (Full ML Churn Prediction)...")
    task = Task(
        name="Analyze Customer Data and Predict Churn",
        description=(
            "MULTIMODAL TASK: Read the provided Excel file 'Customer Data' containing customer metrics. "
            "Analyze the data to identify churn risk patterns. "
            "Create a NEW Excel file called 'churn_predictions.xlsx' with these sheets: "
            "1) 'Analysis' sheet with the original data plus a new 'Churn_Risk' column (High/Medium/Low) "
            "2) 'Summary' sheet with risk distribution statistics. "
            "Also save a markdown report explaining your analysis methodology and key findings."
        ),
        status=TaskStatus.PENDING,
        input_resource_ids=[input_resource.id],
    )
    print(f"  ✓ Task created: {task.name}")
    print("  ✓ Task requires: Excel output + Markdown report")
    print()

    # 4. Create Worker with Tools
    print("Step 4: Creating Worker with GDPEval tools...")

    resource_manager = ResourceFileManager()
    gdpeval_tools = ToolFactory.create_gdpeval_tools(resource_manager=resource_manager)
    print(f"  ✓ Created {len(gdpeval_tools)} tools")

    worker_config = AIAgentConfig(
        agent_id="test_worker",
        agent_type="ai",
        system_prompt="You are a data analyst. Use the read_excel tool to analyze spreadsheets.",
        model_name="gpt-4o-mini",
        agent_description="Test worker",
        agent_capabilities=["data analysis"],
        use_multimodal_resources=True,  # This is the key config!
        max_turns=30,  # Allow more turns for complex tasks
        enable_execution_tracing=True,  # Enable trace capture!
    )

    worker = AIAgent(
        config=worker_config,
        tools=gdpeval_tools,
    )

    # Set communication service after creation
    communication_service = CommunicationService()
    worker.communication_service = communication_service

    print(f"  ✓ Worker created: {worker.config.agent_id}")
    print(f"  ✓ Multimodal mode: {worker.config.use_multimodal_resources}")
    print()

    # 5. Execute Task
    print("Step 5: Executing task...")
    print("-" * 80)

    try:
        result = await worker.execute_task(task, [input_resource])

        print()
        print("=" * 80)
        print("✅ EXECUTION COMPLETE")
        print("=" * 80)
        print(f"Success: {result.success}")
        print(f"Output resources: {len(result.output_resources)}")
        print(f"Total cost: ${result.actual_cost:.4f}")
        print(f"Execution time: {result.execution_time_seconds:.2f}s")
        print(f"Tokens used: {result.tokens_used}")
        print()

        if result.output_resources:
            print(f"✅ Created {len(result.output_resources)} output resources:")
            print()
            for i, res in enumerate(result.output_resources, 1):
                print(f"Output {i}: {res.name}")
                print(f"  Description: {res.description}")
                print(f"  Path: {res.file_path}")
                print(f"  Type: {res.mime_type}")
                print(f"  Size: {res.size_bytes} bytes")

                # Check if file exists
                if res.file_path and Path(res.file_path).exists():
                    print("  ✓ File exists on disk")

                    # Preview based on type
                    if res.is_spreadsheet:
                        try:
                            import pandas as pd

                            xls = pd.ExcelFile(res.file_path)
                            print(
                                f"  📊 Excel file with {len(xls.sheet_names)} sheets:"
                            )
                            for sheet in xls.sheet_names:
                                df = pd.read_excel(xls, sheet_name=sheet)
                                print(
                                    f"    - {sheet}: {len(df)} rows × {len(df.columns)} columns"
                                )
                        except Exception as e:
                            print(f"  Could not read Excel: {e}")

                    elif res.is_text_format:
                        try:
                            text = res.load_text()
                            lines = text.split("\n")
                            print(f"  📝 Text file with {len(lines)} lines")
                            print("  Preview (first 400 chars):")
                            print(f"  {text[:400]}...")
                        except Exception as e:
                            print(f"  Could not preview: {e}")
                else:
                    print("  ⚠️ File not found on disk!")

                print()

        if result.error_message:
            print(f"❌ Error: {result.error_message}")

        # Save execution log
        import json

        log_file = Path(tempfile.gettempdir()) / "test_worker_excel_log.json"
        with open(log_file, "w") as f:
            json.dump(result.model_dump(mode="json"), f, indent=2)
        print(f"📄 Execution log saved to: {log_file}")

        # Check if execution trace was captured
        if result.execution_trace:
            print()
            print("✅ EXECUTION TRACE CAPTURED!")
            print(f"  Total LLM turns: {result.execution_trace.total_turns}")
            print(f"  Total tokens: {result.execution_trace.total_tokens}")
            print(f"  Input tokens: {result.execution_trace.total_input_tokens}")
            print(f"  Output tokens: {result.execution_trace.total_output_tokens}")
            print(f"  Total tool calls: {result.execution_trace.total_tool_calls}")
            print(f"  Duration: {result.execution_trace.duration_seconds:.2f}s")
            print()
            print("  Turn-by-turn breakdown:")
            for turn in result.execution_trace.model_turns:
                print(
                    f"    Turn {turn.turn_index + 1}: {turn.input_tokens} in, {turn.output_tokens} out, {len(turn.tool_calls)} tools"
                )
                for tool_call in turn.tool_calls:
                    print(f"      - {tool_call.tool_name}")
        else:
            print()
            print("⚠️  No execution trace captured (tracing may not be enabled)")
        print()

    except Exception as e:
        print()
        print("=" * 80)
        print("❌ EXECUTION FAILED")
        print("=" * 80)
        print(f"Error: {e}")
        import traceback

        traceback.print_exc()


if __name__ == "__main__":
    asyncio.run(test_worker_excel_file())
